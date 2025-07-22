"""
This is a Python class file to read and write data from an Excel file using the openpyxl library.
The class provides methods to get row and column counts, read data from specific cells, and write 
data to specific cells.
"""

from openpyxl import load_workbook

class ExcelFunctions:
    
    # Initializes the ExcelFunctions object
    def __init__(self, file_name, sheet_name):
        self.file = file_name  # Path to the Excel file
        self.sheet = sheet_name  # Name of the sheet within the Excel file

    # Method to count the total number of rows of the Excel file
    def row_count(self):
        workbook = load_workbook(self.file)  # Load the Excel workbook
        sheet = workbook[self.sheet]  # Access the specified sheet
        return sheet.max_row  # Return the total number of rows in the sheet

    # Method to count the total number of columns of the Excel file
    def column_count(self):
        workbook = load_workbook(self.file)  # Load the Excel workbook
        sheet = workbook[self.sheet]  # Access the specified sheet
        return sheet.max_column  # Return the total number of columns in the sheet
    
    # Method to read data from a specific cell in the Excel sheet
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)  # Load the Excel workbook
        sheet = workbook[self.sheet]  # Access the specified sheet
        return sheet.cell(row=row_number, column=column_number).value  # Return the value of the specified cell

    # Method to write data to a specific cell in the Excel sheet. 
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)  # Load the Excel workbook
        sheet = workbook[self.sheet]  # Access the specified sheet
        sheet.cell(row=row_number, column=column_number).value = data  # Write the data to the specified cell
        workbook.save(self.file)  # Save the changes to the Excel file